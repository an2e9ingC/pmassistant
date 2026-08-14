from __future__ import annotations
from datetime import date, datetime, timedelta
from typing import Optional

from sqlalchemy.orm import Session

from backend.models.zentao import CachedProject, CachedExecution, CachedTask, PmaProduct
from backend.models.bug import CachedBug
from backend.models.delivery import DeliveryRecord
from backend.models.task import WorkLog, Task
from backend.models.bug import BugWorkLog, PmaBug
from backend.models.local import LocalUser
from backend.models.wecom import WeComCheckin


def get_project_summary(db: Session) -> dict:
    """Overall project status summary."""
    projects = db.query(CachedProject).all()
    total = len(projects)
    active = sum(1 for p in projects if p.status in ("doing", "wait"))
    done = sum(1 for p in projects if p.status in ("done", "closed"))
    blocked = sum(1 for p in projects if p.status == "suspended")

    type_all = {}
    for p in projects:
        pt = p.project_type or "RD"
        type_all[pt] = type_all.get(pt, 0) + 1

    return {
        "total": total, "active": active, "done": done, "blocked": blocked,
        "type_all": type_all,
    }


def get_weekly_report(db: Session, project_id: Optional[int] = None) -> dict:
    """Generate a weekly progress report."""
    today = date.today()
    week_ago = today - timedelta(days=7)

    # Completed tasks this week
    tasks_q = db.query(CachedTask)
    if project_id:
        tasks_q = tasks_q.filter(CachedTask.project_id == project_id)
    recent_done_tasks = tasks_q.filter(
        CachedTask.status == "done",
        CachedTask.finished_date >= week_ago,
    ).count()

    # New bugs this week
    bugs_q = db.query(CachedBug)
    if project_id:
        bugs_q = bugs_q.filter(CachedBug.project_id == project_id)
    new_bugs = bugs_q.filter(CachedBug.opened_date >= week_ago).count()
    resolved_bugs = bugs_q.filter(
        CachedBug.resolved_date >= week_ago,
    ).count()

    # Active stages
    execs_q = db.query(CachedExecution)
    if project_id:
        execs_q = execs_q.filter(CachedExecution.project_id == project_id)
    active_stages = execs_q.filter(CachedExecution.status == "doing").all()

    # Delivery this week
    del_q = db.query(DeliveryRecord)
    if project_id:
        del_q = del_q.filter(DeliveryRecord.project_id == project_id)
    week_deliveries = del_q.filter(DeliveryRecord.delivery_date >= week_ago).all()
    week_delivery_qty = sum(r.quantity or 0 for r in week_deliveries)

    return {
        "period": f"{week_ago} ~ {today}",
        "tasks_completed": recent_done_tasks,
        "new_bugs": new_bugs,
        "resolved_bugs": resolved_bugs,
        "active_stages": [{"name": s.name, "project_id": s.project_id} for s in active_stages],
        "delivery_quantity": week_delivery_qty,
        "generated_at": str(today),
    }


def get_monthly_report(db: Session) -> dict:
    """Generate a monthly project report."""
    today = date.today()
    month_start = today.replace(day=1)

    projects = db.query(CachedProject).all()
    summary = get_project_summary(db)

    # Monthly task completion
    done_tasks = db.query(CachedTask).filter(
        CachedTask.status == "done",
        CachedTask.finished_date >= month_start,
    ).count()

    # Monthly bugs
    new_bugs = db.query(CachedBug).filter(
        CachedBug.opened_date >= month_start,
    ).count()
    resolved_bugs = db.query(CachedBug).filter(
        CachedBug.resolved_date >= month_start,
    ).count()

    # Monthly deliveries
    month_deliveries = db.query(DeliveryRecord).filter(
        DeliveryRecord.delivery_date >= month_start,
    ).all()
    month_delivery_qty = sum(r.quantity or 0 for r in month_deliveries)

    # Per-project breakdown
    project_details = []
    for p in projects:
        p_tasks_done = db.query(CachedTask).filter(
            CachedTask.project_id == p.id,
            CachedTask.status == "done",
        ).count()
        p_tasks_total = db.query(CachedTask).filter(
            CachedTask.project_id == p.id,
        ).count()
        project_details.append({
            "id": p.id, "name": p.name, "code": p.code,
            "status": p.status, "progress": p.progress,
            "tasks_done": p_tasks_done,
            "tasks_total": p_tasks_total,
        })

    return {
        "period": f"{month_start} ~ {today}",
        "summary": summary,
        "tasks_completed_this_month": done_tasks,
        "new_bugs_this_month": new_bugs,
        "resolved_bugs_this_month": resolved_bugs,
        "delivery_quantity_this_month": month_delivery_qty,
        "projects": project_details,
        "generated_at": str(today),
    }


def _parse_date_range(date_from: Optional[str], date_to: Optional[str]) -> tuple:
    """Parse optional date range strings. Returns (from_date, to_date)."""
    from_date = datetime.strptime(date_from[:10], "%Y-%m-%d").date() if date_from else date.today().replace(day=1)
    to_date = datetime.strptime(date_to[:10], "%Y-%m-%d").date() if date_to else date.today()
    return from_date, to_date


def _get_effective_hours(worklog) -> float:
    """Get effective hours from worklog: calculated_hours or hours."""
    return float(worklog.calculated_hours or worklog.hours or 0)


def _sum_checkin_hours(db: Session, wuids: list, from_date: date, to_date: date) -> dict:
    """打卡总工时（wecom_userid → hours）：按 (user, date) 取 checkin/approval 的 work_hours 最大值后求和。

    与 wecom_service.get_checkin_calendar 的「每日 max(checkin, approval)」口径一致，
    避免同日既有打卡又有审批时被重复叠加（用户中心与人力报表因此不一致）。
    """
    if not wuids:
        return {}
    rows = db.query(
        WeComCheckin.user_id, WeComCheckin.date,
        WeComCheckin.source, WeComCheckin.work_hours,
    ).filter(
        WeComCheckin.user_id.in_(wuids),
        WeComCheckin.date >= from_date,
        WeComCheckin.date <= to_date,
    ).all()
    per_day = {}  # (wuid, date) -> {"checkin": max, "approval": max}
    for uid, d, src, hrs in rows:
        m = per_day.setdefault((uid, str(d)), {"checkin": 0.0, "approval": 0.0})
        if src == "checkin":
            m["checkin"] = max(m["checkin"], float(hrs or 0.0))
        elif src == "approval":
            m["approval"] = max(m["approval"], float(hrs or 0.0))
    totals = {}
    for (uid, _), m in per_day.items():
        totals[uid] = totals.get(uid, 0.0) + max(m["checkin"], m["approval"])
    return totals


def get_manpower_report(
    db: Session,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    project_id: Optional[int] = None,
) -> dict:
    """Multi-dimensional manpower report."""
    from_date, to_date = _parse_date_range(date_from, date_to)

    # Fetch task worklogs in range
    twl_q = db.query(WorkLog).filter(
        WorkLog.date >= from_date,
        WorkLog.date <= to_date,
    )
    if user_id:
        twl_q = twl_q.filter(WorkLog.user_id == user_id)
    if project_id:
        task_ids = db.query(Task.id).filter(Task.project_id == project_id).all()
        twl_q = twl_q.filter(WorkLog.task_id.in_([t[0] for t in task_ids])) if task_ids else twl_q.filter(WorkLog.task_id == -1)
    task_wls = twl_q.all()

    # Fetch bug worklogs in range
    bwl_q = db.query(BugWorkLog).filter(
        BugWorkLog.date >= from_date,
        BugWorkLog.date <= to_date,
    )
    if user_id:
        bwl_q = bwl_q.filter(BugWorkLog.user_id == user_id)
    bug_wls = bwl_q.all()

    # Build lookup maps
    task_ids = {w.task_id for w in task_wls}
    task_map = {t.id: t for t in db.query(Task).filter(Task.id.in_(task_ids)).all()} if task_ids else {}
    bug_ids = {w.bug_id for w in bug_wls}
    bug_map = {b.id: b for b in db.query(PmaBug).filter(PmaBug.id.in_(bug_ids)).all()} if bug_ids else {}
    proj_ids = {t.project_id for t in task_map.values() if t and t.project_id}
    proj_ids.update(b.project_id for b in bug_map.values() if b and b.project_id)
    proj_map = {}
    prod_map = {}
    if proj_ids:
        for p in db.query(CachedProject).filter(CachedProject.id.in_(proj_ids)).all():
            proj_map[p.id] = p
        for p in db.query(PmaProduct).filter(PmaProduct.id.in_(proj_ids)).all():
            prod_map[p.id] = p
    uid_set = {w.user_id for w in task_wls} | {w.user_id for w in bug_wls}
    user_map = {u.id: u for u in db.query(LocalUser).filter(LocalUser.id.in_(uid_set)).all()} if uid_set else {}

    # ── by_project ──
    by_project = {}
    for w in task_wls:
        t = task_map.get(w.task_id)
        if not t or not t.project_id:
            continue
        pid = t.project_id
        if pid not in by_project:
            proj = proj_map.get(pid) or prod_map.get(pid)
            by_project[pid] = {
                "project_id": pid, "project_code": getattr(proj, "code", "") or "",
                "project_name": getattr(proj, "name", "") or "",
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
                "users": {}, "tasks": {}, "bugs": {},
            }
        h = _get_effective_hours(w)
        by_project[pid]["total_hours"] += h
        by_project[pid]["total_percentage"] += w.percentage or 0
        by_project[pid]["count"] += 1
        by_project[pid]["tasks"][w.task_id] = by_project[pid]["tasks"].get(w.task_id, 0) + h
        uid = str(w.user_id)
        by_project[pid]["users"][uid] = by_project[pid]["users"].get(uid, 0) + h

    for w in bug_wls:
        b = bug_map.get(w.bug_id)
        if not b or not b.project_id:
            continue
        pid = b.project_id
        if pid not in by_project:
            proj = proj_map.get(pid) or prod_map.get(pid)
            by_project[pid] = {
                "project_id": pid, "project_code": getattr(proj, "code", "") or "",
                "project_name": getattr(proj, "name", "") or "",
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
                "users": {}, "tasks": {}, "bugs": {},
            }
        h = _get_effective_hours(w)
        by_project[pid]["total_hours"] += h
        by_project[pid]["total_percentage"] += w.percentage or 0
        by_project[pid]["count"] += 1
        by_project[pid]["bugs"][w.bug_id] = by_project[pid]["bugs"].get(w.bug_id, 0) + h
        uid = str(w.user_id)
        by_project[pid]["users"][uid] = by_project[pid]["users"].get(uid, 0) + h

    # ── by_user ──
    by_user = {}
    for w in task_wls:
        uid = w.user_id
        t = task_map.get(w.task_id)
        if uid not in by_user:
            u = user_map.get(uid)
            by_user[uid] = {
                "user_id": uid, "display_name": (u.display_name or u.username) if u else "?",
                "username": u.username if u else "?",
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
                "projects": {},
            }
        h = _get_effective_hours(w)
        by_user[uid]["total_hours"] += h
        by_user[uid]["total_percentage"] += w.percentage or 0
        by_user[uid]["count"] += 1
        pid = str(t.project_id) if t and t.project_id else "?"
        by_user[uid]["projects"][pid] = by_user[uid]["projects"].get(pid, 0) + h
    for w in bug_wls:
        uid = w.user_id
        b = bug_map.get(w.bug_id)
        if uid not in by_user:
            u = user_map.get(uid)
            by_user[uid] = {
                "user_id": uid, "display_name": (u.display_name or u.username) if u else "?",
                "username": u.username if u else "?",
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
                "projects": {},
            }
        h = _get_effective_hours(w)
        by_user[uid]["total_hours"] += h
        by_user[uid]["total_percentage"] += w.percentage or 0
        by_user[uid]["count"] += 1
        pid = str(b.project_id) if b and b.project_id else "?"
        by_user[uid]["projects"][pid] = by_user[uid]["projects"].get(pid, 0) + h

    # ── by_product ──
    by_product = {}
    for w in task_wls:
        t = task_map.get(w.task_id)
        if not t or not t.project_id:
            continue
        proj = prod_map.get(t.project_id)
        if not proj:
            continue
        pid = t.project_id
        if pid not in by_product:
            by_product[pid] = {
                "product_id": pid, "product_code": getattr(proj, "code", "") or "",
                "product_name": getattr(proj, "name", "") or "",
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
            }
        h = _get_effective_hours(w)
        by_product[pid]["total_hours"] += h
        by_product[pid]["total_percentage"] += w.percentage or 0
        by_product[pid]["count"] += 1

    # ── Summary ──
    total_hours = sum(p["total_hours"] for p in by_project.values())
    total_count = sum(p["count"] for p in by_project.values())
    person_count = len(by_user)

    # Format by_project for API response
    formatted_projects = []
    for pid, pdata in sorted(by_project.items(), key=lambda x: -x[1]["total_hours"]):
        users_list = []
        for uid_str, hrs in sorted(pdata["users"].items(), key=lambda x: -x[1]):
            u = user_map.get(int(uid_str))
            users_list.append({
                "user_id": int(uid_str),
                "display_name": (u.display_name or u.username) if u else "?",
                "hours": round(hrs, 1),
            })
        tasks_list = [{"task_id": tid, "title": (task_map.get(int(tid)).title if task_map.get(int(tid)) else "?"),
                       "hours": round(hrs, 1)} for tid, hrs in sorted(pdata["tasks"].items(), key=lambda x: -x[1])]
        bugs_list = [{"bug_id": int(bid), "hours": round(hrs, 1)} for bid, hrs in sorted(pdata.get("bugs", {}).items(), key=lambda x: -x[1])]
        formatted_projects.append({
            "project_id": pdata["project_id"], "project_code": pdata["project_code"],
            "project_name": pdata["project_name"],
            "total_hours": round(pdata["total_hours"], 1),
            "total_share": round(pdata["total_hours"] / total_hours * 100, 1) if total_hours > 0 else 0.0,
            "percentage_avg": round(pdata["total_percentage"] / max(pdata["count"], 1), 1),
            "users": users_list, "tasks": tasks_list[:20], "bugs": bugs_list[:20],
        })

    # ── 打卡工时（企业微信）：覆盖所有绑定 wecom_userid 的本地用户，wecom_userid → 打卡工时 ──
    checkin_map = {}
    wecom_user_by_id = {}
    wecom_users = db.query(LocalUser).filter(
        LocalUser.wecom_userid.isnot(None),
        LocalUser.wecom_userid != "",
    ).all()
    if wecom_users:
        wecom_user_by_id = {u.wecom_userid: u for u in wecom_users}
        wuids = list(wecom_user_by_id.keys())
        checkin_map = _sum_checkin_hours(db, wuids, from_date, to_date)

    # 补充：有打卡、但无 PMA 工时记录的人员，使其也出现在报表中
    for wuid, u in wecom_user_by_id.items():
        if u.id not in by_user and wuid in checkin_map:
            by_user[u.id] = {
                "user_id": u.id,
                "display_name": u.display_name or u.username,
                "username": u.username,
                "total_hours": 0.0, "total_percentage": 0.0, "count": 0,
                "projects": {},
            }
            user_map[u.id] = u
    person_count = len(by_user)

    formatted_users = []
    for uid, udata in sorted(by_user.items(), key=lambda x: -x[1]["total_hours"]):
        u = user_map.get(uid)
        checkin_hours = checkin_map.get(u.wecom_userid, 0.0) if u and u.wecom_userid else 0.0
        pma_hours = round(udata["total_hours"], 1)
        ratio = round(pma_hours / checkin_hours * 100, 1) if checkin_hours > 0 else None
        formatted_users.append({
            "user_id": udata["user_id"], "display_name": udata["display_name"],
            "username": udata["username"],
            "total_hours": pma_hours,
            "checkin_hours": round(checkin_hours, 1),
            "pma_hours": pma_hours,
            "ratio": ratio,  # 记录/打卡占比（%），无打卡为 None
            "percentage_avg": round(udata["total_percentage"] / max(udata["count"], 1), 1),
            "project_count": len(udata["projects"]),
        })

    formatted_products = []
    for pid, pdata in sorted(by_product.items(), key=lambda x: -x[1]["total_hours"]):
        formatted_products.append({
            "product_id": pdata["product_id"], "product_code": pdata["product_code"],
            "product_name": pdata["product_name"],
            "total_hours": round(pdata["total_hours"], 1),
            "percentage_avg": round(pdata["total_percentage"] / max(pdata["count"], 1), 1),
        })

    return {
        "summary": {
            "total_hours": round(total_hours, 1),
            "person_count": person_count,
            "project_count": len(by_project),
            "product_count": len(by_product),
        },
        "by_project": formatted_projects,
        "by_user": formatted_users,
        "by_product": formatted_products,
        "period": {"from": str(from_date), "to": str(to_date)},
    }


def get_user_manpower_detail(
    db: Session,
    user_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> dict:
    """Per-user manpower detail: project breakdown (pie) + daily breakdown."""
    from_date, to_date = _parse_date_range(date_from, date_to)

    twls = db.query(WorkLog).filter(
        WorkLog.user_id == user_id,
        WorkLog.date >= from_date,
        WorkLog.date <= to_date,
    ).all()
    bws = db.query(BugWorkLog).filter(
        BugWorkLog.user_id == user_id,
        BugWorkLog.date >= from_date,
        BugWorkLog.date <= to_date,
    ).all()

    task_ids = {w.task_id for w in twls}
    bug_ids = {w.bug_id for w in bws}
    task_map = {t.id: t for t in db.query(Task).filter(Task.id.in_(task_ids)).all()} if task_ids else {}
    bug_map = {b.id: b for b in db.query(PmaBug).filter(PmaBug.id.in_(bug_ids)).all()} if bug_ids else {}
    proj_ids = {t.project_id for t in task_map.values() if t and t.project_id}
    proj_ids.update(b.project_id for b in bug_map.values() if b and b.project_id)
    proj_map = {}
    prod_map = {}
    if proj_ids:
        for p in db.query(CachedProject).filter(CachedProject.id.in_(proj_ids)).all():
            proj_map[p.id] = p
        for p in db.query(PmaProduct).filter(PmaProduct.id.in_(proj_ids)).all():
            prod_map[p.id] = p

    def _proj_meta(pid):
        proj = proj_map.get(pid) or prod_map.get(pid)
        return {
            "project_id": pid,
            "project_code": getattr(proj, "code", "") or "",
            "project_name": getattr(proj, "name", "") or "",
        }

    # ── Project breakdown ──
    proj_hours = {}
    for w in twls:
        t = task_map.get(w.task_id)
        if not t or not t.project_id:
            continue
        proj_hours[t.project_id] = proj_hours.get(t.project_id, 0.0) + _get_effective_hours(w)
    for w in bws:
        b = bug_map.get(w.bug_id)
        if not b or not b.project_id:
            continue
        proj_hours[b.project_id] = proj_hours.get(b.project_id, 0.0) + _get_effective_hours(w)

    total_hours = sum(proj_hours.values())
    projects = []
    for pid, hrs in sorted(proj_hours.items(), key=lambda x: -x[1]):
        meta = _proj_meta(pid)
        meta["hours"] = round(hrs, 1)
        meta["percentage"] = round(hrs / total_hours * 100, 1) if total_hours > 0 else 0.0
        projects.append(meta)

    # ── Daily breakdown ──
    daily_map = {}
    for w in twls:
        t = task_map.get(w.task_id)
        pid = t.project_id if t and t.project_id else None
        daily_map.setdefault(str(w.date), {}).setdefault(pid, 0.0)
        daily_map[str(w.date)][pid] += _get_effective_hours(w)
    for w in bws:
        b = bug_map.get(w.bug_id)
        pid = b.project_id if b and b.project_id else None
        daily_map.setdefault(str(w.date), {}).setdefault(pid, 0.0)
        daily_map[str(w.date)][pid] += _get_effective_hours(w)

    daily = []
    for d in sorted(daily_map.keys(), reverse=True):
        day_hours = sum(daily_map[d].values())
        proj_list = []
        for pid, hrs in sorted(daily_map[d].items(), key=lambda x: -x[1]):
            meta = _proj_meta(pid) if pid else {"project_id": None, "project_code": "", "project_name": "其他"}
            meta["hours"] = round(hrs, 1)
            meta["percentage"] = round(hrs / day_hours * 100, 1) if day_hours > 0 else 0.0
            proj_list.append(meta)
        daily.append({"date": d, "hours": round(day_hours, 1), "projects": proj_list})

    # ── Checkin hours ──
    user = db.query(LocalUser).filter(LocalUser.id == user_id).first()
    checkin_hours = 0.0
    if user and user.wecom_userid:
        checkin_hours = _sum_checkin_hours(db, [user.wecom_userid], from_date, to_date).get(user.wecom_userid, 0.0)

    pma_hours = round(total_hours, 1)
    ratio = round(pma_hours / checkin_hours * 100, 1) if checkin_hours > 0 else None

    return {
        "summary": {
            "user_id": user_id,
            "display_name": (user.display_name or user.username) if user else "?",
            "checkin_hours": round(checkin_hours, 1),
            "pma_hours": pma_hours,
            "ratio": ratio,
        },
        "projects": projects,
        "daily": daily,
    }


def export_manpower_excel(
    db: Session,
    output,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    project_id: Optional[int] = None,
):
    """Export manpower report as Excel (.xlsx) using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    from_date, to_date = _parse_date_range(date_from, date_to)
    data = get_manpower_report(db, str(from_date), str(to_date), user_id, project_id)

    wb = openpyxl.Workbook()

    # 数据查询（供各 sheet 复用）
    twls = db.query(WorkLog).filter(WorkLog.date >= from_date, WorkLog.date <= to_date)
    if user_id: twls = twls.filter(WorkLog.user_id == user_id)
    twls = twls.order_by(WorkLog.date.desc()).all()
    bws = db.query(BugWorkLog).filter(BugWorkLog.date >= from_date, BugWorkLog.date <= to_date)
    if user_id: bws = bws.filter(BugWorkLog.user_id == user_id)
    bws = bws.order_by(BugWorkLog.date.desc()).all()
    task_ids = {w.task_id for w in twls}
    bug_ids = {w.bug_id for w in bws}
    task_map = {t.id: t for t in db.query(Task).filter(Task.id.in_(task_ids)).all()} if task_ids else {}
    bug_map = {b.id: b for b in db.query(PmaBug).filter(PmaBug.id.in_(bug_ids)).all()} if bug_ids else {}
    uid_set = {w.user_id for w in twls} | {w.user_id for w in bws}
    user_map = {u.id: u for u in db.query(LocalUser).filter(LocalUser.id.in_(uid_set)).all()} if uid_set else {}
    proj_map = {}
    for p in db.query(CachedProject).all(): proj_map[p.id] = (p.code, p.name)
    for p in db.query(PmaProduct).all(): proj_map[p.id] = (p.code, p.name)

    # Sheet 1: 按人员-项目工时占比（财务核算每人各项目人力投入）
    ws1 = wb.active
    ws1.title = "按人员-项目占比"
    ws1.append(["人员", "项目编号", "项目名", "工时(h)", "占该人总工时比例(%)"])

    user_proj = {}
    for w in twls:
        t = task_map.get(w.task_id)
        if not t or not t.project_id:
            continue
        uid = w.user_id
        if uid not in user_proj:
            user_proj[uid] = {}
        user_proj[uid][t.project_id] = user_proj[uid].get(t.project_id, 0.0) + _get_effective_hours(w)
    for w in bws:
        b = bug_map.get(w.bug_id)
        if not b or not b.project_id:
            continue
        uid = w.user_id
        if uid not in user_proj:
            user_proj[uid] = {}
        user_proj[uid][b.project_id] = user_proj[uid].get(b.project_id, 0.0) + _get_effective_hours(w)

    for uid in sorted(user_proj.keys(), key=lambda u: -sum(user_proj[u].values())):
        u = user_map.get(uid)
        uname = (u.display_name or u.username) if u else "?"
        total = sum(user_proj[uid].values())
        for pid, hrs in sorted(user_proj[uid].items(), key=lambda x: -x[1]):
            code, name = proj_map.get(pid, ("", ""))
            pct = round(hrs / total * 100, 1) if total > 0 else 0.0
            ws1.append([uname, code, name, round(hrs, 1), pct])

    # Sheet 1b: 按人员-项目工时占企微打卡工时百分比
    checkin_by_uid = {u["user_id"]: u.get("checkin_hours", 0.0) for u in data.get("by_user", [])}
    ws1b = wb.create_sheet("按人员-项目占比(打卡)")
    ws1b.append(["人员", "项目编号", "项目名", "工时(h)", "占企微打卡工时比例(%)"])
    for uid in sorted(user_proj.keys(), key=lambda u: -sum(user_proj[u].values())):
        u = user_map.get(uid)
        uname = (u.display_name or u.username) if u else "?"
        checkin = checkin_by_uid.get(uid, 0.0)
        for pid, hrs in sorted(user_proj[uid].items(), key=lambda x: -x[1]):
            code, name = proj_map.get(pid, ("", ""))
            pct = round(hrs / checkin * 100, 1) if checkin > 0 else 0.0
            ws1b.append([uname, code, name, round(hrs, 1), pct])

    # Sheet 2: 工时明细
    ws2 = wb.create_sheet("工时明细")
    ws2.append(["人员", "项目编号", "项目名", "产品编号", "产品名", "日期", "工时占比(%)", "计算工时(h)", "工作内容", "来源"])
    for w in twls:
        t = task_map.get(w.task_id)
        proj_info = proj_map.get(t.project_id, ("", "")) if t and t.project_id else ("", "")
        u = user_map.get(w.user_id)
        ws2.append([
            (u.display_name or u.username) if u else "?", proj_info[0], proj_info[1], "", "",
            str(w.date), w.percentage or "", w.calculated_hours or w.hours or "",
            w.description or "", "task",
        ])
    for w in bws:
        b = bug_map.get(w.bug_id)
        proj_info = proj_map.get(b.project_id, ("", "")) if b and b.project_id else ("", "")
        u = user_map.get(w.user_id)
        ws2.append([
            (u.display_name or u.username) if u else "?", proj_info[0], proj_info[1], "", "",
            str(w.date), w.percentage or "", w.calculated_hours or w.hours or "",
            w.description or "", "bug",
        ])

    # Sheet 3: 按项目汇总
    ws3 = wb.create_sheet("按项目汇总")
    ws3.append(["项目编号", "项目名", "总工时(h)", "人均占比(%)"])
    for p in data.get("by_project", []):
        ws3.append([p["project_code"], p["project_name"], p["total_hours"], p["percentage_avg"]])

    # Sheet 4: 按人员汇总
    ws4 = wb.create_sheet("按人员汇总")
    ws4.append(["人员", "用户名", "打卡工时(h)", "PMA记录工时(h)", "记录/打卡占比(%)", "涉及项目数"])
    for u in data.get("by_user", []):
        ws4.append([
            u["display_name"], u["username"],
            u.get("checkin_hours", ""), u["pma_hours"],
            u["ratio"] if u.get("ratio") is not None else "",
            u["project_count"],
        ])

    wb.save(output)
